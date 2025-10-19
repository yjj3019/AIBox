#!/usr/bin/env python3.11
# -*- coding: utf-8 -*-

# cve 파일이 들어있는 txt 파일을 입력받아 그에 상응하는 cve 리포틀 만드는 소스

import os
import json
import requests
import sys
import re
import time
import argparse
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed
import logging

# [사용자 요청] 서버 측 Excel 생성을 위해 openpyxl 라이브러리 추가
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
except ImportError:
    openpyxl = None

# --- 기본 설정 ---
ANALYSIS_PERIOD_DAYS = 180
CVE_DATA_FILE = '/data/iso/AIBox/meta/cve_data.json'
OUTPUT_DIR = '/data/iso/AIBox/output'
REPORT_FILE_PATH = os.path.join(OUTPUT_DIR, 'rhel_vulnerability_report.html')
MAX_WORKERS = 10


# --- API 엔드포인트 ---
LOCAL_CVE_URL = 'http://127.0.0.1:5000/AIBox/cve/{cve_id}.json'
REDHAT_CVE_URL = 'https://access.redhat.com/hydra/rest/securitydata/cve/{cve_id}.json'
LLM_ANALYZE_URL = 'http://127.0.0.1:5000/AIBox/api/cve/analyze'

# --- 필터링할 RHEL 제품 목록 (정규표현식 사용) ---
TARGET_PRODUCT_PATTERNS = [
    re.compile(r"^Red Hat Enterprise Linux 7$"),
    re.compile(r"^Red Hat Enterprise Linux 7 Extended Lifecycle Support$"),
    re.compile(r"^Red Hat Enterprise Linux 8$"),
    re.compile(r"^Red Hat Enterprise Linux 8\.\d+ Extended Update Support$"),
    re.compile(r"^Red Hat Enterprise Linux 8\.\d+ Extended Update Support Long-Life Add-On$"),
    re.compile(r"^Red Hat Enterprise Linux 8\.\d+ Update Services for SAP Solutions$"),
    re.compile(r"^Red Hat Enterprise Linux 9$"),
    re.compile(r"^Red Hat Enterprise Linux 9\.\d+ Extended Update Support$"),
    re.compile(r"^Red Hat Enterprise Linux 9\.\d+ Extended Update Support Long-Life Add-On$"),
    re.compile(r"^Red Hat Enterprise Linux 9\.\d+ Update Services for SAP Solutions$"),
    re.compile(r"^Red Hat Enterprise Linux 10$"),
    re.compile(r"^Red Hat Enterprise Linux 10\.\d+ Extended Update Support$"),
    re.compile(r"^Red Hat Enterprise Linux 10\.\d+ Extended Update Support Long-Life Add-On$"),
    re.compile(r"^Red Hat Enterprise Linux 10\.\d+ Update Services for SAP Solutions$"),
    re.compile(r"^Red Hat Enterprise Linux \d+\.\d+ for SAP Solutions$"),
    re.compile(r"^Red Hat Enterprise Linux \d+\.\d+ Update Services for SAP Solutions$")
]

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', stream=sys.stdout)

def collect_cve_ids_from_custom_file(file_path):
    """--cve-file 옵션으로 제공된 파일에서 CVE ID 목록을 읽습니다."""
    logging.info(f"사용자 지정 파일에서 CVE 목록 수집을 시작합니다: {file_path}")
    if not os.path.exists(file_path):
        logging.error(f"오류: CVE 파일 '{file_path}'를 찾을 수 없습니다.")
        return []

    cve_ids = set()
    cve_pattern = re.compile(r'^(CVE-\d{4}-\d{4,})$')
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            for line in f:
                clean_line = line.strip()
                if match := cve_pattern.match(clean_line):
                    cve_ids.add(match.group(1))
    except Exception as e:
        logging.error(f"'{file_path}' 파일 처리 중 오류 발생: {e}")
        return []
    
    logging.info(f"파일에서 {len(cve_ids)}개의 유효한 CVE ID를 발견했습니다.")
    return sorted(list(cve_ids), reverse=True)

def collect_cve_ids_from_file():
    logging.info(f"데이터 파일에서 CVE 목록 수집을 시작합니다: {CVE_DATA_FILE}")
    try:
        with open(CVE_DATA_FILE, 'r', encoding='utf-8') as f:
            all_cves = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError) as e:
        logging.error(f"오류: CVE 데이터 파일을 읽을 수 없습니다. {e}")
        return []

    cve_ids = set()
    start_date = datetime.now() - timedelta(days=ANALYSIS_PERIOD_DAYS)
    for cve in all_cves:
        try:
            if 'public_date' in cve:
                public_date = datetime.strptime(cve['public_date'].split('T')[0], '%Y-%m-%d')
                if public_date >= start_date:
                    cve_id = cve.get('name') or cve.get('CVE')
                    if cve_id: cve_ids.add(cve_id)
        except (ValueError, TypeError):
            continue
    logging.info(f"분석 기간 내 {len(cve_ids)}개의 고유 CVE를 발견했습니다.")
    return sorted(list(cve_ids), reverse=True)

def fetch_cve_details(cve_id):
    """[수정] 재시도 로직이 추가된 CVE 상세 정보 조회 함수"""
    urls_to_try = [
        {"url": LOCAL_CVE_URL.format(cve_id=cve_id), "source": "로컬 서버"},
        {"url": REDHAT_CVE_URL.format(cve_id=cve_id), "source": "Red Hat API"}
    ]
    max_retries = 3

    for source_info in urls_to_try:
        url, source = source_info["url"], source_info["source"]
        for attempt in range(max_retries):
            try:
                response = requests.get(url, timeout=30)
                if response.status_code == 200:
                    # [사용자 요청] 재시도 후 성공 시 로그를 명확히 남깁니다.
                    if attempt > 0: logging.info(f" -> [{cve_id}] Successfully connected to {source} on attempt {attempt + 1}/{max_retries}.")
                    data = response.json()
                    if 'name' not in data: data['name'] = cve_id
                    return data, cve_id
                if response.status_code == 404: break # 404는 재시도 불필요
                response.raise_for_status()
            except requests.RequestException as e:
                logging.warning(f" -> [{cve_id}] {source} 연결 오류 (시도 {attempt + 1}): {e}")
                if attempt < max_retries - 1: time.sleep(1)

    logging.error(f"[{cve_id}] 모든 소스에서 정보 조회 실패.")
    return None, cve_id

def process_cve_data(cve_data, skip_filtering=False):
    if not cve_data: return None
    cve_id = cve_data.get('name', 'N/A')

    # [사용자 요청] 필터링 규칙 설명을 위한 변수 초기화
    filter_reason = "선정"
    filtered_releases = {}

    all_releases = cve_data.get('affected_release', [])
    if not all_releases and not skip_filtering:
        filter_reason = "영향받는 제품 없음"
    else:
        for release in all_releases:
            product_name = release.get('product_name', 'Unknown Product')
            # --cve-file 옵션 사용 시 필터링을 건너뜁니다.
            if not skip_filtering:
                is_target_product = any(pattern.match(product_name) for pattern in TARGET_PRODUCT_PATTERNS)
                if not is_target_product:
                    filter_reason = "분석 대상 제품 아님"
                    continue

                package = release.get('package', '')
                if 'kpatch-patch' in package or 'kernel-rt' in package:
                    filter_reason = "라이브 커널 패치 또는 실시간 커널 제외"
                    continue

                is_not_affected = any(
                    s.get('product_name') == product_name and s.get('fix_state') == 'Not affected'
                    for s in cve_data.get('package_state', [])
                )
                if is_not_affected:
                    filter_reason = "패치 불필요(Not Affected)"
                    continue

            advisory = release.get('advisory', 'N/A')
            if product_name not in filtered_releases: filtered_releases[product_name] = {}
            if advisory not in filtered_releases[product_name]: filtered_releases[product_name][advisory] = []
            filtered_releases[product_name][advisory].append(release.get('package', ''))

        if not filtered_releases and filter_reason == "선정":
            filter_reason = "관련 릴리즈 없음"

    # 필터링되어 릴리즈 정보가 없다면, 필터링 이유를 포함하여 반환
    if not filtered_releases and not skip_filtering:
        return {'cve_id': cve_id, 'filter_reason': filter_reason}

    bugzilla_info = cve_data.get('bugzilla', {})

    # [사용자 요청] 'affected_release' 필드를 기반으로 패키지 이름 추출
    affected_package_names = set()
    if isinstance(cve_data.get('affected_release'), list):
        for release in cve_data['affected_release']:
            # [BUG FIX] TARGET_PRODUCT_PATTERNS에 해당하는 제품의 패키지만 고려합니다.
            product_name = release.get('product_name', 'Unknown Product')
            if not skip_filtering and not any(pattern.match(product_name) for pattern in TARGET_PRODUCT_PATTERNS):
                continue

            full_pkg_name = release.get('package')
            if not full_pkg_name: continue
            
            match = re.match(r'^(.*)-[^-]+-[^-]+$', full_pkg_name)
            if match:
                base_pkg_name = match.group(1).rsplit('-', 1)[0] if ':' in match.group(1) else match.group(1)
                affected_package_names.add(base_pkg_name)

    return {
        'cve_id': cve_id,
        'public_date': cve_data.get('public_date', 'N/A').split('T')[0],
        'severity': cve_data.get('threat_severity', 'N/A'),
        'cvss3_score': cve_data.get('cvss3', {}).get('cvss3_base_score', 'N/A'),
        'summary_en': cve_data.get('statement') or " ".join(cve_data.get('details', [])),
        'affected_releases': filtered_releases,
        'bugzilla_id': bugzilla_info.get('id'),
        'bugzilla_url': bugzilla_info.get('url'),
        'cwe_id': cve_data.get('cwe'),
        'affected_package_names': sorted(list(affected_package_names)),
        'filter_reason': filter_reason # [사용자 요청] 필터링 결과 추가
    }

def get_korean_summaries_with_llm(cve_list):
    """
    [NEW] security.py 로직 적용: 모든 CVE를 한번에 LLM에 요청하여 번역 안정성 극대화
    """
    if not cve_list:
        return []
        
    logging.info(f"LLM 국문 요약 생성을 시작합니다: 총 {len(cve_list)}개 CVE (단일 배치 요청)")

    analyzed_data = {}

    def analyze_single_cve(cve):
        """[NEW] security.py 로직을 참고하여 단일 CVE에 대한 심층 분석을 요청하는 함수"""
        cve_id = cve['cve_id']
        # AI에게 전달할 데이터 형식에 맞게 입력 데이터를 가공
        cve_for_prompt = {
            "cve_id": cve_id,
            "severity": cve.get('severity', 'N/A'), # type: ignore
            "cvss_score": cve.get('cvss3_score', 'N/A'),
            "summary": cve.get('summary_en', '')
        }

        # security.py의 전문가 프롬프트를 적용
        prompt = {
            "task": "research_and_summarize_cve_deep_dive_in_korean",
            "prompt": ( # type: ignore
                "You are a top-tier cybersecurity expert specializing in Red Hat Enterprise Linux (RHEL). "
                "Your mission is to conduct a deep-dive analysis of the provided CVE. "
                "You MUST perform a web search to gather the latest threat intelligence. "
                "All analysis must be in PERFECT, NATURAL KOREAN."
                "\n\n[Analysis Guidelines & Web Search Requirements]"
                "\n1.  **Threat Intelligence Gathering (Web Search)**: Search for CISA KEV (Known Exploited Vulnerabilities) and PoC (Proof-of-Concept) code availability (e.g., Exploit-DB, GitHub)."
                "\n2.  **Detailed Analysis**: Based on the gathered intelligence, provide:"
                "\n    - **threat_tags**: Identify threat types like 'RCE', 'Privilege Escalation'. If it's in CISA KEV, you MUST include the 'Exploited in wild' tag. If a PoC is public, add 'PoC Available'."
                "\n    - **concise_summary**: A 2-3 sentence summary in Korean for technical experts." # type: ignore
                "\n\n[Output Format]"
                "\nReturn ONLY a single, valid JSON object with keys: 'threat_tags', 'concise_summary'."
            ),
            "cve_data": cve_for_prompt, # type: ignore
            "model_selector": "deep_dive" # [수정] 서버가 reasoning_model을 선택하도록 키워드 추가 # type: ignore
        }
        max_retries = 3
        for attempt in range(max_retries):
            try:
                response = requests.post(LLM_ANALYZE_URL, json=prompt, timeout=180, headers={'Content-Type': 'application/json'}) # type: ignore
                response.raise_for_status()
                result = response.json()
                if all(k in result for k in ['threat_tags', 'concise_summary']):
                    return cve_id, result
                else:
                    logging.warning(f"[{cve_id}] LLM이 예상치 못한 형식의 JSON을 반환했습니다: {result}")
                    return cve_id, None # 형식 오류도 실패로 간주
            except requests.RequestException as e:
                logging.warning(f"[{cve_id}] LLM API 요청 오류 (시도 {attempt + 1}): {e}")
                if attempt < max_retries - 1: time.sleep(1)
            except json.JSONDecodeError:
                logging.warning(f"[{cve_id}] LLM 응답이 JSON이 아님: {response.text[:100]}")
                break # JSON 형식이 아니면 재시도 의미 없음

        logging.error(f"[{cve_id}] LLM 분석 최종 실패.")
        return cve_id, None

    # ThreadPoolExecutor를 사용하여 병렬로 각 CVE 분석 요청
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_to_cve = {executor.submit(analyze_single_cve, cve): cve['cve_id'] for cve in cve_list}
        
        processed_count = 0
        for future in as_completed(future_to_cve):
            cve_id = future_to_cve[future]
            try:
                _, analysis_result = future.result()
                if analysis_result:
                    analyzed_data[cve_id] = analysis_result
                else:
                    analyzed_data[cve_id] = {"concise_summary": "[AI 요약 생성 실패]", "threat_tags": []}
            except Exception as exc:
                logging.error(f"'{cve_id}' 처리 중 예외 발생: {exc}")
                analyzed_data[cve_id] = {"concise_summary": "[AI 요약 중 오류 발생]", "threat_tags": []}
            
            processed_count += 1
            logging.info(f"-> 진행률: {processed_count}/{len(cve_list)} ({cve_id})")
            
            # [사용자 요청] 서버 부하 조절을 위해 10개 처리마다 5초 대기
            if processed_count % 10 == 0 and processed_count < len(cve_list):
                logging.info("   -> 서버 부하 조절을 위해 5초간 대기합니다...")
                time.sleep(5)

    # 원본 cve_list에 번역 결과를 매칭하여 추가
    for cve in cve_list:
        cve['analysis'] = analyzed_data.get(cve['cve_id'], {})
    
    logging.info("LLM 국문 요약 생성 완료.")
    return cve_list

def make_request_with_retry(method, url, max_retries=3, **kwargs):
    """[신규] 재시도 로직이 포함된 범용 요청 함수"""
    for attempt in range(max_retries):
        try:
            response = requests.request(method, url, **kwargs)
            response.raise_for_status()
            # [사용자 요청] 재시도 후 성공 시 로그를 명확히 남깁니다.
            if attempt > 0:
                logging.info(f"Successfully connected on attempt {attempt + 1}/{max_retries} for request to {url}")
            return response
        except requests.RequestException as e:
            logging.warning(f"Request failed (Attempt {attempt + 1}/{max_retries}): {url}, Error: {e}")
            if attempt < max_retries - 1:
                wait_time = 2 ** attempt
                time.sleep(wait_time)
            else:
                logging.error(f"Request failed after {max_retries} retries: {url}")
                return None
    return None





def generate_interactive_html_report(cve_list, custom_report_path=None):
    logging.info("HTML 리포트 생성을 시작합니다.")

    table_rows_html = ""
    # [개선] 클라이언트 측 Excel 생성 로직 제거. 서버에서 직접 생성합니다.

    for cve in cve_list:
        flat_releases = []
        # [요청사항] CVE 정보가 없는 경우 처리
        if cve.get('error'):
            table_rows_html += f"""
            <tr>
                <td><a href="https://access.redhat.com/security/cve/{cve['cve_id']}" target="_blank">{cve['cve_id']}</a></td>
                <td colspan="7" style="text-align: center; color: #777;">{cve['error']}</td>
            </tr>
            """
            continue

        # [사용자 요청] 필터링된 CVE에 대한 설명 추가
        if cve.get('filter_reason') and cve.get('filter_reason') != '선정':
            table_rows_html += f"""
            <tr>
                <td><a href="https://access.redhat.com/security/cve/{cve['cve_id']}" target="_blank">{cve['cve_id']}</a></td>
                <td colspan="7" style="text-align: center; color: #777;">{cve['filter_reason']}</td>
            </tr>
            """
            continue
        if cve.get('affected_releases'):
            for product, advisories in cve.get('affected_releases', {}).items():
                for advisory, packages in advisories.items():
                    # [사용자 요청] 패키지 이름에서 'idm:', 'redis:' 같은 접두사 제거
                    cleaned_packages = []
                    for pkg in packages:
                        cleaned_packages.append(pkg.split(':', 1)[-1])
                    package_list_html = "<br>".join(cleaned_packages)
                    flat_releases.append({
                        'product': product, 'advisory': advisory, 'packages': package_list_html
                    })
        
        analysis = cve.get("analysis", {})
        # [BUG FIX] AI 분석 결과가 예상치 못한 형식(예: 문자열)일 경우를 대비한 안정성 강화
        if isinstance(analysis, dict):
            summary = analysis.get("concise_summary", "요약 없음").replace('\n', '<br>')
            threat_tags = analysis.get("threat_tags", [])
        else:
            summary = "AI 분석 결과를 처리하는 중 오류가 발생했습니다."
            threat_tags = []

        # [신규] 위협 태그 HTML 생성
        tags_html = ""
        if threat_tags:
            for tag in threat_tags:
                tag_class = "tag-exploited" if "Exploited" in tag else "tag-threat"
                tags_html += f'<span class="threat-tag {tag_class}">{tag}</span>'
        
        # [수정] 요약 정보에서 '분석 근거(selection_reason)' 제거
        # [사용자 요청] Bugzilla, CWE, Package 정보 추가
        reference_links_html = ""
        if cve.get("bugzilla_id") and cve.get("bugzilla_url"):
            reference_links_html += f'<span>RHBZ: <a href="{cve["bugzilla_url"]}" target="_blank">{cve["bugzilla_id"]}</a></span>'
        
        if cve.get("cwe_id"):
            cwe_id_clean = cve["cwe_id"].split('(')[0].strip() # "CWE-1220(Generation of Error Message...)" -> "CWE-1220"
            cwe_num = re.search(r'\d+', cwe_id_clean)
            if cwe_num:
                cwe_url = f"https://cwe.mitre.org/data/definitions/{cwe_num.group()}.html"
                if reference_links_html: reference_links_html += ' / '
                reference_links_html += f'<span>CWE: <a href="{cwe_url}" target="_blank">{cwe_id_clean}</a></span>'

        if cve.get("affected_package_names"):
            if reference_links_html: reference_links_html += ' / '
            reference_links_html += f'<span>package : {", ".join(cve["affected_package_names"])}</span>'

        summary_html_block = f"""
        <div class="summary-content">
            <div class="summary-tags">{tags_html}</div>
            <p class="main-summary">{summary}</p>
            <div class="reference-links">{reference_links_html}</div>
        </div>
        """
        
        # [BUG FIX] 테이블 레이아웃이 깨지는 문제를 해결하기 위해 HTML 생성 로직을 재구성합니다.
        # 1. rowspan에 사용할 행의 개수를 계산합니다.
        if not flat_releases:
            flat_releases.append({'product': '-', 'advisory': '-', 'packages': '-'})
        num_rows_for_cve = len(flat_releases)

        # 2. 각 CVE에 대한 행을 생성합니다.
        for i, release_info in enumerate(flat_releases):
            table_rows_html += '<tr>'
            # 첫 번째 행에만 rowspan을 사용하여 공통 정보를 병합합니다.
            if i == 0:
                severity = cve.get("severity", "N/A")
                cve_id = cve["cve_id"]
                cve_url = f"https://access.redhat.com/security/cve/{cve_id}"
                cve_link_html = f'<a href="{cve_url}" target="_blank">{cve_id}</a>'
                table_rows_html += f'<td rowspan="{num_rows_for_cve}">{cve_link_html}</td>'
                table_rows_html += f'<td rowspan="{num_rows_for_cve}">{cve["public_date"]}</td>'
                table_rows_html += f'<td rowspan="{num_rows_for_cve}">{severity}</td>'
                table_rows_html += f'<td rowspan="{num_rows_for_cve}">{cve["cvss3_score"]}</td>'
                table_rows_html += f'<td rowspan="{num_rows_for_cve}" class="summary-cell">{summary_html_block}</td>'
            
            # 각 행에 고유한 릴리즈 정보를 추가합니다.
            advisory = release_info["advisory"]
            advisory_html = f'<a href="https://access.redhat.com/errata/{advisory}" target="_blank">{advisory}</a>' if advisory != 'N/A' and advisory.startswith('RHSA') else advisory
            table_rows_html += f'<td>{release_info["product"]}</td>'
            table_rows_html += f'<td>{advisory_html}</td>'
            table_rows_html += f'<td class="package-cell">{release_info["packages"]}</td>'
            table_rows_html += '</tr>'
    
    html_template = f"""
    <!DOCTYPE html>
    <html lang="ko">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Red Hat 취약점 분석 리포트</title>
        <link rel="preconnect" href="https://fonts.googleapis.com">
        <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
        <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;700&display=swap" rel="stylesheet">
        <style>
            body {{
                font-family: 'Noto Sans KR', sans-serif;
                margin: 0;
                padding: 20px;
                background-color: #f4f7f9;
                color: #333;
            }}
            .container {{
                max-width: 98%;
                margin: 0 auto;
                background-color: #ffffff;
                border-radius: 12px;
                box-shadow: 0 8px 24px rgba(149, 157, 165, 0.2);
                padding: 30px;
            }}
            .report-header {{
                text-align: center;
                margin-bottom: 30px;
                border-bottom: 2px solid #e0e0e0;
                padding-bottom: 20px;
            }}
            .report-header h1 {{
                color: #c92127; /* Red Hat Red */
                font-size: 2.2em;
                font-weight: 700;
                margin: 0;
            }}
            .report-header p {{
                color: #666;
                font-size: 1.1em;
                margin-top: 5px;
            }}
            .controls {{
                margin-bottom: 20px;
                text-align: right;
            }}
            #export-btn {{
                background-color: #0088ce;
                color: white;
                border: none;
                padding: 12px 24px;
                font-size: 16px;
                font-weight: 700;
                cursor: pointer;
                border-radius: 8px;
                transition: background-color 0.3s, box-shadow 0.3s;
                box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            }}
            #export-btn:hover {{
                background-color: #006da7;
                box-shadow: 0 4px 8px rgba(0,0,0,0.15);
            }}
            #report-table {{
                width: 100%;
                border-collapse: collapse;
                border: 1px solid #e0e0e0;
                border-radius: 10px;
                overflow: hidden; /* For border-radius on table */
            }}
            #report-table th, #report-table td {{
                padding: 15px;
                text-align: center;
                vertical-align: middle;
                border-bottom: 1px solid #e0e0e0;
            }}
            #report-table th {{
                background-color: #333a40;
                color: white;
                font-weight: 700;
                font-size: 1.05em;
            }}
            #report-table tbody tr:hover {{
                background-color: #eef7ff;
            }}
            #report-table td a {{
                color: #0066cc;
                text-decoration: none;
                font-weight: 700;
            }}
            #report-table td a:hover {{
                text-decoration: underline;
            }}
            .summary-cell {{
                text-align: left;
                white-space: pre-wrap;
                word-break: break-word;
            }}
            .package-cell {{ text-align: left; }}
            .summary-tags {{ margin-bottom: 10px; }}
            .threat-tag {{
                display: inline-block; padding: .2em .6em; margin-right: .5rem;
                margin-bottom: .3rem; font-size: .8rem; font-weight: 700;
                color: #fff; border-radius: 4px;
            }}
            .tag-exploited {{ background-color: #c92127; }}
            .tag-threat {{ background-color: #f57c00; }}
            .summary-content .main-summary {{ margin: 0 0 10px 0; }}
            .reference-links {{
                margin-top: 12px;
                padding-top: 10px;
                border-top: 1px dashed #ccc;
                font-size: 0.9em;
                color: #555;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="report-header">
                <h1>Red Hat Enterprise Linux 취약점 분석 리포트</h1>
                <p>AI 기반 자동 요약 및 번역 | 최종 업데이트: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            </div>            
            <div class="controls"><button id="export-btn">Excel로 저장</button></div>
            <table id="report-table">
                <thead><tr>
                    <th style="width: 12%;">CVE ID</th><th style="width: 8%;">발행 일자</th><th style="width: 8%;">심각도</th><th style="width: 8%;">CVSSv3</th>
                    <th style="width: 24%;">취약점 요약</th><th style="width: 15%;">영향받는 제품</th><th style="width: 10%;">RHSA ID</th><th style="width: 15%;">패치된 패키지</th>
                </tr></thead>
                <tbody>{table_rows_html}</tbody>
            </table>
        </div>
        <script src="https://cdnjs.cloudflare.com/ajax/libs/xlsx/0.18.5/xlsx.full.min.js"></script>
        <script>
            document.getElementById('export-btn').addEventListener('click', function() {{
                // [개선] 서버의 새 엔드포인트를 호출하여 Excel 파일을 다운로드합니다.
                window.location.href = '/AIBox/api/cve/export-excel';
            }});
        </script>
    </body>
    </html>
    """
    report_path = custom_report_path if custom_report_path else REPORT_FILE_PATH

    try:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write(html_template)
        logging.info(f"성공: HTML 리포트가 '{report_path}'에 저장되었습니다.")
    except IOError as e:
        logging.error(f"오류: HTML 리포트 파일 저장 실패: {e}")

def create_excel_report(cve_list):
    """[신규] openpyxl을 사용하여 서버 측에서 직접 Excel 파일을 생성합니다."""
    if not openpyxl:
        logging.error("Excel 생성 오류: 'openpyxl' 라이브러리가 설치되지 않았습니다.")
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RHEL 취약점 리포트"

    headers = [
        "CVE ID", "발행 일자", "심각도", "CVSSv3", "취약점 요약", 
        "영향받는 패키지 (전체)", "영향받는 제품", "RHSA ID", "패치된 패키지"
    ]
    ws.append(headers)

    # 헤더 스타일 설정
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="333a40", end_color="333a40", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # 데이터 추가
    for cve in cve_list:
        analysis = cve.get("analysis", {})
        flat_releases = []
        if cve.get('affected_releases'):
            for product, advisories in cve.get('affected_releases', {}).items():
                for advisory, packages in advisories.items():
                    flat_releases.append({
                        'product': product, 
                        'advisory': advisory, 
                        'packages': ", ".join(pkg.split(':', 1)[-1] for pkg in packages)
                    })
        if not flat_releases:
            flat_releases.append({'product': '-', 'advisory': '-', 'packages': '-'})

        # 여러 릴리즈 정보를 한 셀에 줄 바꿈으로 통합
        products = "\n".join([fr['product'] for fr in flat_releases])
        advisories = "\n".join([fr['advisory'] for fr in flat_releases])
        packages = "\n".join([fr['packages'] for fr in flat_releases])

        row_data = [
            cve["cve_id"], cve["public_date"], cve.get("severity", "N/A"), cve["cvss3_score"],
            analysis.get("concise_summary", "요약 없음"), ", ".join(cve.get("affected_package_names", [])),
            products, advisories, packages
        ]
        ws.append(row_data)
        
        # 하이퍼링크 추가
        cve_cell = ws.cell(row=ws.max_row, column=1)
        cve_cell.hyperlink = f"https://access.redhat.com/security/cve/{cve['cve_id']}"
        cve_cell.font = Font(color="0000FF", underline="single")

    # 컬럼 너비 및 스타일 조정
    column_widths = [20, 15, 12, 10, 60, 40, 40, 20, 50]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=i, max_col=i):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    return wb

def main():
    logging.info("===== Red Hat 취약점 분석 스크립트(v29, security.py 로직 적용)를 시작합니다. =====")
    
    parser = argparse.ArgumentParser(description="Red Hat CVE 분석 리포트 생성기")
    parser.add_argument('--cve-file', help='분석할 CVE ID 목록이 포함된 텍스트 파일 경로')
    args = parser.parse_args()

    if args.cve_file:
        cve_ids_to_fetch = collect_cve_ids_from_custom_file(args.cve_file)
    else:
        cve_ids_to_fetch = collect_cve_ids_from_file()

    if not cve_ids_to_fetch:
        logging.warning("분석할 CVE가 없습니다. 종료합니다.")
        return

    all_cve_details = []
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        f_to_cve = {executor.submit(fetch_cve_details, cid): cid for cid in cve_ids_to_fetch}
        for f in as_completed(f_to_cve):
            data, cve_id = f.result()
            if data:
                all_cve_details.append(data)
            else:
                # [요청사항] CVE 정보 조회 실패 시, 기본 구조를 생성하여 리포트에 포함
                all_cve_details.append({'cve_id': cve_id, 'error': 'CVE 정보 영향 없음(Red Hat)'})
    
    processed_cves = []
    for d in all_cve_details:
        # [사용자 요청] 필터링 설명을 위해, 조회 실패 시에도 cve_id와 에러 메시지를 포함하여 전달
        if d.get('error'):
            processed_cves.append(d)
            continue
        # --cve-file 옵션 사용 시 필터링 건너뛰기
        processed = process_cve_data(d, skip_filtering=bool(args.cve_file))
        # [사용자 요청] process_cve_data가 필터링되어 None을 반환하는 대신, 이유가 담긴 dict를 반환하도록 수정됨
        if processed:
            processed_cves.append(processed)

    if not processed_cves:
        logging.warning("필터링 후 리포트에 포함할 CVE가 없습니다.")
        report_path = None
        if args.cve_file:
            date_str = datetime.now().strftime('%Y%m%d')
            report_path = os.path.join(OUTPUT_DIR, f'rhel_vulnerability_report_{date_str}.html')
        generate_interactive_html_report([], custom_report_path=report_path)
        return
        
    logging.info(f"데이터 정제 완료. 분석 대상 CVE는 총 {len(processed_cves)}개 입니다.")

    # [NEW] AI 분석 및 국문 요약 생성 (security.py의 안정적인 단일 요청 방식 적용)
    # [사용자 요청] AI 분석 대상은 필터링을 통과한 CVE들로 한정
    cves_for_ai_analysis = [cve for cve in processed_cves if not cve.get('error') and cve.get('filter_reason') == '선정']
    final_cve_list = get_korean_summaries_with_llm(cves_for_ai_analysis)

    # 에러가 발생한 CVE들을 최종 목록에 다시 추가
    final_cve_list.extend([cve for cve in processed_cves if cve.get('error')])
    
    # [개선] 생성된 최종 CVE 목록을 파일로 저장하여 서버가 사용할 수 있도록 함
    try:
        with open(os.path.join(OUTPUT_DIR, 'final_cve_list.json'), 'w', encoding='utf-8') as f:
            json.dump(final_cve_list, f, ensure_ascii=False, indent=2)
        logging.info("최종 CVE 목록을 'final_cve_list.json' 파일로 저장했습니다.")
    except IOError as e:
        logging.error(f"최종 CVE 목록 파일 저장 실패: {e}")

    # [요청사항] --cve-file 사용 시 파일명에 날짜 추가
    report_path = None
    if args.cve_file:
        date_str = datetime.now().strftime('%Y%m%d')
        report_path = os.path.join(OUTPUT_DIR, f'rhel_vulnerability_report_{date_str}.html')

    generate_interactive_html_report(final_cve_list, custom_report_path=report_path)

    logging.info("===== 모든 작업이 완료되었습니다. =====")

if __name__ == '__main__':
    main()
